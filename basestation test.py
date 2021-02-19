import os
import sys
import datetime
import time
import threading
import serial
import RPi.GPIO as GPIO 
from time import sleep    
from datetime import datetime
from xbee import ZigBee
import urllib
import binascii
import socket
import string
import psycopg2 as pg
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT
import json
from gpiozero import CPUTemperature
from time import sleep, strftime, time
from openpyxl import Workbook

wb = Workbook()
ws0 = wb.create_sheet("Temperature data", 0)

active_ws = wb["Temperature data"]
active_ws.cell(row=1, column=1, value="Time")
active_ws.cell(row=1, column=2, value="Temperature")
wb.save('Temperature data.xlsx')
print('saved')


select=set()
enter=1
xbee_reg=0
start_tx_time=0
success_flag=0
xbee_register=0
prev_time=0
counter=1
sent_t = 0
manual = ''
rnum_t = 0
rid_t = ''

#broadcast add 0x000000000000FFFF
dest_addr_long_b=['0x00','0x00','0x00','0X00','0x00','0x00','0xFF','0xFF']
dest_addr_long_b= bytearray( int(x,16) for  x in dest_addr_long_b)

#co ordinator add 0013A200410755E5
dest_addr_long_1=['0x00','0x13','0xA2','0X00','0x41','0x06','0xF5','0x64']
dest_addr_long_1= bytearray( int(x,16) for  x in dest_addr_long_1)   

ser=serial.Serial('/dev/ttyUSB0',9600,timeout=0)
xbee=ZigBee(ser)

class MyDb(object):


    def __init__(self):
        self.conn=pg.connect(
                database = "robot_data",
                user = "nocca",
                password = "admin",
                host = "169.254.29.176",
                port = "5432"
                )       
        self.conn.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
        self.cursor=self.conn.cursor()
        #print('connected')

        
    def db_create_table(self):
            query = """CREATE TABLE IF NOT EXISTS Register_table(robot_number SERIAL NOT NULL UNIQUE , xbee_id CHAR(16) NOT NULL PRIMARY KEY UNIQUE, row_number INT NOT NULL,response_time FLOAT4,route_discovered CHAR(50), timestamp BIGINT NOT NULL);"""
            self.cursor.execute(query)
            #print("table created")

            query = """CREATE TABLE IF NOT EXISTS Monitoring_data(robot_number INT REFERENCES Register_table(robot_number),xbee_id CHAR(16) REFERENCES Register_table(xbee_id), data CHAR(100) NOT NULL, message_id BIGINT NOT NULL, timestamp BIGINT PRIMARY KEY);"""
            self.cursor.execute(query)      #data format is "battey:op_h:robot_status:GPS_data"
            #print("table created")

            query = """CREATE TABLE IF NOT EXISTS Notification_data(robot_number INT REFERENCES Register_table(robot_number), xbee_id CHAR(16) REFERENCES Register_table(xbee_id), status CHAR(50), timestamp BIGINT PRIMARY KEY);"""
            self.cursor.execute(query)      #status can be "dock", "running", "breakdown", "available", "returning"
            #print("table created")

            query = """CREATE TABLE IF NOT EXISTS Command_data(robot_number INT REFERENCES Register_table(robot_number), xbee_id CHAR(16) REFERENCES Register_table(xbee_id), command CHAR(50) NOT NULL, timestamp BIGINT NOT NULL PRIMARY KEY);"""
            self.cursor.execute(query)      #command can be "start", "stop", "dock" and schedule_time is a timestamp
            #print("table created")

            query = """CREATE TABLE IF NOT EXISTS Maintenance_data(robot_number INT REFERENCES Register_table(robot_number), xbee_id CHAR(16) REFERENCES Register_table(xbee_id), M_data CHAR(200), timestamp BIGINT NOT NULL PRIMARY KEY);"""
            self.cursor.execute(query)      #all replacements are timestamps
            #print("table created")
            
    def db_insert_R_table(self, robot_id_s):     
            query="SELECT xbee_id FROM Register_table"
            self.cursor.execute(query)
            x=self.cursor.fetchall() 
            repeat_reg = 0
            print(x)      
            for n in x:
                for m in n:
                     print(m)
                     if (m==robot_id_s):
                        repeat_reg = 1
            if repeat_reg!=1:
                query= """INSERT INTO Register_table(xbee_id, row_number,response_time,route_discovered,timestamp) VALUES (%s,%s,%s,%s,%s)"""
                data=(robot_id_s, 1,1,'{"route":"0000"}', timestamp_curr)
                self.cursor.execute(query,data)
                print('done')
                self.conn.commit()
                print("data inserted to r table")
            else:
                print("copies")

    def db_insert_M_table(self, robot_id_s, motor_rep, proxy_rep, batt_rep, brush_rep):
            query="SELECT robot_number FROM Register_table WHERE (xbee_id)='{}'".format(robot_id_s)
            self.cursor.execute(query)            
            y=self.cursor.fetchone()      
            data = {
                    "motor_replacement":motor_rep,
                    "proxy_replacement":proxy_rep, 
                    "battery_replacement":batt_rep,
                    "brush_replacement":brush_rep
                    }
            data_json=json.dumps(data)
            query= "INSERT INTO Maintenance_data(xbee_id, robot_number, timestamp, M_data) VALUES (%s,%s,%s,%s)"
            data= (robot_id_s, y[0], timestamp_curr, data_json)            
            self.cursor.execute(query,data)
            self.conn.commit()
            print("data inserted to m table")            

    def db_insert_N_table(self, robot_id_s, status):
            query="SELECT robot_number FROM Register_table WHERE (xbee_id)='{}'".format(robot_id_s)
            self.cursor.execute(query)            
            y=self.cursor.fetchone()    
            data = {
                    "status":status
                    }
            data_json=json.dumps(data)    
            query= "INSERT INTO  Notification_data(xbee_id, robot_number, status,timestamp) VALUES (%s,%s,%s,%s)"
            data= (robot_id_s, y[0], data_json, timestamp_curr)                      
            self.cursor.execute(query,data)
            self.conn.commit()
            print("data inserted to n table")    

    def db_insert_C_table(self, robot_id_s, command):
            query="SELECT robot_number FROM Register_table WHERE (xbee_id)='{}'".format(robot_id_s)
            self.cursor.execute(query)            
            y=self.cursor.fetchone()   
            data = {
                    "command":command
                    }
            data_json=json.dumps(data)     
            query= "INSERT INTO  Command_data(xbee_id, robot_number, command, timestamp) VALUES (%s,%s,%s,%s)"
            data= (robot_id_s, y[0], data_json, timestamp_curr)                      
            self.cursor.execute(query,data)
            self.conn.commit()
            print("data inserted to c table")

    def db_insert_Mo_table(self, robot_id_s,batter_s,op_h,status,gps_s,msg_id):
            query="SELECT MAX(message_id) FROM Monitoring_data WHERE xbee_id='{}'".format(robot_id_s)
            self.cursor.execute(query)            
            m_id = self.cursor.fetchone()
            print(m_id[0])
            if (bool(m_id[0]==msg_id)==False):
                query="SELECT robot_number FROM Register_table WHERE (xbee_id)='{}'".format(robot_id_s)
                self.cursor.execute(query)            
                y=self.cursor.fetchone() 
                data = {
                        "battey":batter_s,
                        "operation_hours":op_h,
                        "robot_status":status,
                        "GPS_data":gps_s
                        }
                data_json=json.dumps(data)  
                print(data_json)     
                query= "INSERT INTO  Monitoring_data(xbee_id, robot_number, data, message_id, timestamp) VALUES (%s,%s,%s,%s,%s)"
                data= (robot_id_s, y[0], data_json, msg_id, timestamp_curr)                      
                self.cursor.execute(query,data)
                self.conn.commit()
                print("data inserted to Mo table")
            else:
                print("copies")
    def ip_thread(self, en):        
            while en==1:       
                global manual    
                manual=raw_input("User input:")
                if manual=="check":
                        robot=raw_input("Robot number:")
                        query="SELECT xbee_id FROM Register_table WHERE robot_number = '{}'".format(robot)
                        self.cursor.execute(query)
                        y=self.cursor.fetchone()
                        robot_id_s=binascii.a2b_hex(y[0])
                        self.db_insert_C_table(y[0], "route check")
                        xbee.tx(dest_addr_long=robot_id_s,frame_id='\x01',options='\x08',data="onetimed1")
                if manual=="ok":
                        global select
                        select=set()
                if manual=="start":
                        query="SELECT MAX(robot_number) FROM Register_table"
                        self.cursor.execute(query)            
                        y=self.cursor.fetchone()
                        for n in range(y[0]):
                                global rnum_t 
                                rnum_t = n+1
                                global success_flag     
                                success_flag=0              
                                retry=0
                                query="SELECT xbee_id FROM Register_table WHERE robot_number = '{}'".format(n+1)
                                self.cursor.execute(query)
                                y=self.cursor.fetchone()
                                global rid_t
                                rid_t = y[0]
                                robot_id_s=binascii.a2b_hex(y[0])
                                self.db_insert_C_table(y[0], "start")
                                while success_flag==0 and (retry)!=4:
                                        global sent_t
                                        sent_t=timestamp_curr
                                        xbee.tx(dest_addr_long=robot_id_s,frame_id='\x01',data="startmesh")
                                        retry+=1
                                        sleep(3)   #waiting for 3 seconds
                                if retry==4 and success_flag==0:
                                        self.db_insert_N_table(y[0],'unable to connect')
                                        retry=0
                sleep(2)
                if manual=="exit":
                        en=0
obj=MyDb()                
i = threading.Thread(target=obj.ip_thread, args=(enter,))
i.start() 
while True:
    cpu = CPUTemperature()
    temp=cpu.temperature

    x1=datetime.now()
    x1=x1.strftime("%H:%M:%S")   
    obj=MyDb()
    query = "CREATE DATABASE IF NOT EXISTS robot_data"
    #obj.cursor.execute(query)
    #print("database created")
    obj.db_create_table()
    #print("tables created")

    timestamp_curr = time()
    if ((timestamp_curr) - (prev_time) >= 20):
        active_ws = wb["Temperature data"]
        counter += 1
        active_ws.cell(row=counter, column=1, value=timestamp_curr)
        active_ws.cell(row=counter, column=2, value=temp)
        wb.save('Temperature data.xlsx')
        prev_time = timestamp_curr

    

    def Zigbee_data(data): 
         print(data)   
         global manual

         xbee_data=data['id']       
         xbee_data=xbee_data.decode()
         data_id=xbee_data
         print(data_id)

         if data_id=='rx':                 #this is for 
                xbee_data=data['source_addr_long']
                robot_id=binascii.b2a_hex(xbee_data)
                robot_id_s=robot_id.decode()
                print(robot_id_s)
                #xbee_data=binascii.a2b_hex(robot_id_s)         
               
                xbee_data=data['rf_data']       
                xbee_data=xbee_data.decode()
                split_data=xbee_data.split(":")
                request_type=split_data[0]
                print(request_type)
         
         if data_id == 'route_information':                
                global select
                responder=data['responder_addr']                
                responder_1=binascii.b2a_hex(responder)
                responder_1=responder_1.decode()                
                query="SELECT robot_number FROM Register_table WHERE xbee_id='{}'".format(responder_1)
                obj.cursor.execute(query)            
                y=obj.cursor.fetchone()                
                if responder_1=='0013a200410755e3':                
                        y=[0,]                
                select.add(y[0])                

                dest=data['dest_addr']
                dest_1=binascii.b2a_hex(dest)
                dest_1=dest_1.decode()
                query="SELECT robot_number FROM Register_table WHERE xbee_id='{}'".format(dest_1)
                obj.cursor.execute(query)            
                y=obj.cursor.fetchone()
                select.add(y[0])              

                source=data['source_addr']
                source_1=binascii.b2a_hex(source)
                source_1=source_1.decode()
                query="SELECT robot_number FROM Register_table WHERE xbee_id='{}'".format(source_1)
                obj.cursor.execute(query)            
                y=obj.cursor.fetchone()
                if responder_1=='0013a200410755e3':
                        y=[0,]
                select.add(y[0])                

                rece=data['receiver_addr']
                rece_1=binascii.b2a_hex(rece)
                rece_1=rece_1.decode()
                query="SELECT robot_number FROM Register_table WHERE xbee_id='{}'".format(rece_1)
                obj.cursor.execute(query)            
                y=obj.cursor.fetchone()
                select.add(y[0])            

                route=list(select)
                route.sort()               
                route1=''.join(str(ele) for ele in route)
                              
                query="UPDATE Register_table SET route_discovered='{a}' WHERE xbee_id='{b}'".format(a=route1, b=dest_1)
                obj.cursor.execute(query)
                print('route updated')

         if data_id=="tx_status" and manual=="start":                
                receive_t=timestamp_curr
                global sent_t
                response_time=(receive_t)-(sent_t) 
                d_status=data['deliver_status']
                d_status=binascii.b2a_hex(d_status)
                d_status=d_status.decode() 

                xbee_data=data['dest_addr']
                robot_id=binascii.b2a_hex(xbee_data)
                robot_id_s=robot_id.decode()

                if d_status=='00':
                        print("///////")
                        global success_flag
                        success_flag=1
                        obj.db_insert_N_table(rid_t,'running') 
                        print(response_time)                        
                        query="UPDATE Register_table SET response_time='{}' WHERE robot_number='{}'".format(response_time,rnum_t)
                        obj.cursor.execute(query)                            
                        
                
         if request_type=="SQ":
                battery_s=int(split_data[1])    #data format is "battey:op_h:robot_status:GPS_data:message_id"
                oph_s=int(split_data[2])
                robot_status_s=split_data[3]
                gps_s=split_data[4]
                message_id = int(split_data[5])               
                #obj.db_insert_N_table(robot_id_s,temp)
                print("ok")
                obj.db_insert_Mo_table(robot_id_s,battery_s,oph_s,robot_status_s,gps_s,message_id)           

         if request_type=="Register":                           
                obj.db_insert_R_table(robot_id_s)
                obj.db_insert_N_table(robot_id_s,'available') 


    xbee=ZigBee(ser,callback=Zigbee_data) 
    xbee.halt()